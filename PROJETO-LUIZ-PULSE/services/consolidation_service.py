import pandas as pd
import openpyxl
import time
import os
from datetime import datetime
from typing import List, Callable
from models.consolidation_model import (
    ConsolidationConfig, ConsolidationResult, 
    ConsolidationStep, FileInfo
)
from services.file_service import FileService
from services.formatting_service import FormattingService, FormattingPriority

class ConsolidationService:
    """Serviço principal para consolidação de dados"""
    
    def __init__(self):
        self.file_service = FileService()
        self.formatting_service = FormattingService()
        self.steps = [
            ConsolidationStep(1, "Validação", "Validando arquivos e configurações", "pending"),
            ConsolidationStep(2, "Preparação", "Preparando arquivos para consolidação", "pending"),
            ConsolidationStep(3, "Leitura", "Lendo dados das planilhas subordinadas", "pending"),
            ConsolidationStep(4, "Consolidação", "Consolidando dados na planilha mestre", "pending"),
            ConsolidationStep(5, "Formatação", "Aplicando formatação das planilhas subordinadas", "pending"),
            ConsolidationStep(6, "Backup e Finalização", "Criando backup e finalizando processo", "pending")
        ]
    
    def consolidate_data(
        self, 
        config: ConsolidationConfig, 
        subordinate_files: List[FileInfo],
        progress_callback: Callable[[int, str, float], None] = None
    ) -> ConsolidationResult:
        """Executa a consolidação completa"""
        
        start_time = time.time()
        errors = []
        total_rows_added = 0
        backup_path = None
        
        try:
            # Step 1: Validação
            self._update_step_status(1, "running", 0.0, progress_callback)
            
            if not self._validate_configuration(config, subordinate_files):
                errors.append("Configuração inválida")
                self._update_step_status(1, "error", 0.0, progress_callback, "Configuração inválida")
                return self._create_result(False, 0, 0, None, time.time() - start_time, errors)
            
            self._update_step_status(1, "completed", 100.0, progress_callback)
            
            # Step 2: Preparação
            self._update_step_status(2, "running", 0.0, progress_callback)
            
            # Apenas validar que o arquivo mestre pode ser lido/escrito
            if not os.access(config.master_file_path, os.R_OK | os.W_OK):
                errors.append("Arquivo mestre não tem permissões adequadas")
                self._update_step_status(2, "error", 0.0, progress_callback, "Permissões inadequadas")
                return self._create_result(False, 0, 0, None, time.time() - start_time, errors)
            
            self._update_step_status(2, "completed", 100.0, progress_callback)
            
            # Step 3: Leitura dos dados
            self._update_step_status(3, "running", 0.0, progress_callback)
            
            master_df = self.file_service.read_excel_data(config.master_file_path, config.master_sheet_name)
            if master_df is None:
                errors.append("Erro ao ler planilha mestre")
                self._update_step_status(3, "error", 0.0, progress_callback, "Erro ao ler planilha mestre")
                return self._create_result(False, 0, 0, backup_path, time.time() - start_time, errors)
            
            subordinate_data = []
            for i, file_info in enumerate(subordinate_files):
                df = self.file_service.read_excel_data(file_info.path, file_info.sheet_names[0])
                if df is not None:
                    subordinate_data.append(df)
                else:
                    errors.append(f"Erro ao ler arquivo: {file_info.name}")
                
                progress = ((i + 1) / len(subordinate_files)) * 100
                self._update_step_status(3, "running", progress, progress_callback)
            
            self._update_step_status(3, "completed", 100.0, progress_callback)
            
            # Step 4: Consolidação direta no arquivo Excel
            self._update_step_status(4, "running", 0.0, progress_callback)
            
            total_rows_added = self._consolidate_excel_direct(config.master_file_path, config.master_sheet_name, subordinate_data, config)
            
            if total_rows_added < 0:
                errors.append("Erro ao consolidar dados no arquivo mestre")
                self._update_step_status(4, "error", 0.0, progress_callback, "Erro na consolidação")
                return self._create_result(False, len(subordinate_files), 0, backup_path, time.time() - start_time, errors)
            
            self._update_step_status(4, "completed", 100.0, progress_callback)
            
            # Step 5: Formatação das planilhas subordinadas
            self._update_step_status(5, "running", 0.0, progress_callback)
            
            # Aplicar formatação das planilhas subordinadas
            subordinate_paths = [f.path for f in subordinate_files]
            formatting_success = self.formatting_service.process_formatting(
                config.master_file_path,
                subordinate_paths,
                strategy='MERGE'
            )
            
            if not formatting_success:
                errors.append("Aviso: Erro ao aplicar formatação das planilhas subordinadas")
            
            self._update_step_status(5, "completed", 100.0, progress_callback)
            
            # Step 6: Backup e Finalização
            self._update_step_status(6, "running", 0.0, progress_callback)
            
            # Criar backup APÓS consolidação bem-sucedida
            if config.backup_enabled:
                backup_path = self.file_service.create_backup(config.master_file_path)
                if not backup_path:
                    errors.append("Aviso: Falha ao criar backup, mas consolidação foi bem-sucedida")
                else:
                    print(f"[ConsolidationService] Backup criado: {backup_path}")
            
            self._update_step_status(6, "completed", 100.0, progress_callback)
            
            execution_time = time.time() - start_time
            return self._create_result(True, len(subordinate_files), total_rows_added, backup_path, execution_time, errors)
            
        except Exception as e:
            errors.append(f"Erro inesperado: {str(e)}")
            execution_time = time.time() - start_time
            return self._create_result(False, 0, 0, backup_path, execution_time, errors)
    
    def _validate_configuration(self, config: ConsolidationConfig, files: List[FileInfo]) -> bool:
        """Valida a configuração da consolidação"""
        if not config.master_file_path or not files:
            return False
        
        # Verifica se o arquivo mestre existe
        master_info = self.file_service.get_file_info(config.master_file_path)
        if not master_info:
            return False
        
        # Verifica se a planilha especificada existe
        if config.master_sheet_name not in master_info.sheet_names:
            return False
        
        return True
    
    def _consolidate_excel_direct(self, master_file_path: str, sheet_name: str, subordinate_dfs: List[pd.DataFrame], config: ConsolidationConfig) -> int:
        """Consolida dados diretamente no arquivo Excel preservando formatação"""
        try:
            print(f"[ConsolidationService] Consolidando {len(subordinate_dfs)} arquivos no arquivo mestre")
            
            # Abrir o arquivo mestre com openpyxl para preservar formatação
            workbook = openpyxl.load_workbook(master_file_path)
            
            if sheet_name not in workbook.sheetnames:
                print(f"[ConsolidationService] Planilha '{sheet_name}' não encontrada")
                return -1
            
            worksheet = workbook[sheet_name]
            
            # Encontrar a última linha com dados
            last_row = worksheet.max_row
            total_rows_added = 0
            
            # Adicionar dados de cada arquivo subordinado
            for i, df in enumerate(subordinate_dfs, 1):
                print(f"[ConsolidationService] Processando arquivo {i}/{len(subordinate_dfs)} - {len(df)} linhas")
                
                # Estratégia de consolidação
                if config.merge_strategy == "append":
                    # Adicionar todas as linhas ao final (exceto cabeçalho se existir)
                    start_row = last_row + 1
                    
                    for row_idx, (_, row) in enumerate(df.iterrows()):
                        for col_idx, value in enumerate(row.values, 1):
                            cell = worksheet.cell(row=start_row + row_idx, column=col_idx)
                            cell.value = value
                    
                    rows_added = len(df)
                    last_row += rows_added
                    total_rows_added += rows_added
                    
                elif config.merge_strategy == "replace":
                    # Substituir dados existentes (implementação simplificada)
                    print(f"[ConsolidationService] Estratégia 'replace' não totalmente implementada")
                    
                elif config.merge_strategy == "update":
                    # Atualizar baseado em colunas-chave (implementação simplificada - append)
                    start_row = last_row + 1
                    
                    for row_idx, (_, row) in enumerate(df.iterrows()):
                        for col_idx, value in enumerate(row.values, 1):
                            cell = worksheet.cell(row=start_row + row_idx, column=col_idx)
                            cell.value = value
                    
                    rows_added = len(df)
                    last_row += rows_added
                    total_rows_added += rows_added
            
            # Salvar o arquivo com formatação preservada
            workbook.save(master_file_path)
            workbook.close()
            
            print(f"[ConsolidationService] Consolidação concluída: {total_rows_added} linhas adicionadas")
            return total_rows_added
            
        except Exception as e:
            print(f"[ConsolidationService] Erro na consolidação direta: {str(e)}")
            return -1
    
    def _merge_dataframes(self, master_df: pd.DataFrame, subordinate_dfs: List[pd.DataFrame], config: ConsolidationConfig) -> pd.DataFrame:
        """Consolida os DataFrames conforme a estratégia configurada (método legado)"""
        result_df = master_df.copy()
        
        for df in subordinate_dfs:
            if config.merge_strategy == "append":
                # Adiciona todas as linhas ao final
                result_df = pd.concat([result_df, df], ignore_index=True)
            elif config.merge_strategy == "replace":
                # Substitui dados existentes (apenas para demonstração)
                result_df = df.copy()
            elif config.merge_strategy == "update":
                # Atualiza dados existentes baseado em colunas-chave
                # Implementação simplificada - adiciona novas linhas
                result_df = pd.concat([result_df, df], ignore_index=True)
        
        return result_df
    
    def _update_step_status(self, step_id: int, status: str, progress: float, callback: Callable = None, error_msg: str = None):
        """Atualiza o status de um passo"""
        for step in self.steps:
            if step.id == step_id:
                step.status = status
                step.progress = progress
                step.error_message = error_msg
                break
        
        if callback:
            callback(step_id, status, progress)
    
    def _create_result(self, success: bool, files_processed: int, rows_added: int, backup_path: str, execution_time: float, errors: List[str]) -> ConsolidationResult:
        """Cria o resultado da consolidação"""
        return ConsolidationResult(
            success=success,
            total_files_processed=files_processed,
            total_rows_added=rows_added,
            backup_path=backup_path,
            execution_time=execution_time,
            errors=errors,
            steps=self.steps.copy()
        )
    
    def get_consolidation_steps(self) -> List[ConsolidationStep]:
        """Retorna os passos da consolidação"""
        return self.steps.copy()