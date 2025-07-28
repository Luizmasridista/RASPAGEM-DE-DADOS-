"""
Serviço de Formatação e Layout
=============================

Este módulo é responsável por transferir formatação, filtros, cores,
cabeçalhos e fórmulas das planilhas subordinadas para a planilha mestre.

Estratégia Modular:
1. FormattingExtractor - Extrai formatação de planilhas subordinadas
2. FormattingApplier - Aplica formatação na planilha mestre
3. FormattingMerger - Gerencia conflitos e prioridades
4. FormattingValidator - Valida compatibilidade de formatação
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, IconSetRule
from openpyxl.worksheet.filters import AutoFilter
from typing import Dict, List, Optional, Any, Tuple
import copy
from dataclasses import dataclass, field
from enum import Enum
import os


class FormattingPriority(Enum):
    """Prioridades para resolução de conflitos de formatação"""
    MASTER = "master"  # Prioridade para formatação da planilha mestre
    SUBORDINATE = "subordinate"  # Prioridade para formatação das subordinadas
    MERGE = "merge"  # Mesclar formatações quando possível


@dataclass
class CellFormatting:
    """Formatação de uma célula"""
    font: Optional[Font] = None
    fill: Optional[PatternFill] = None
    border: Optional[Border] = None
    alignment: Optional[Alignment] = None
    number_format: Optional[str] = None
    formula: Optional[str] = None


@dataclass
class ColumnFormatting:
    """Formatação de uma coluna"""
    width: Optional[float] = None
    hidden: bool = False
    auto_filter: bool = False
    conditional_formatting: List[Any] = field(default_factory=list)


@dataclass
class SheetFormatting:
    """Formatação completa de uma planilha"""
    name: str
    cell_formats: Dict[str, CellFormatting] = field(default_factory=dict)
    column_formats: Dict[str, ColumnFormatting] = field(default_factory=dict)
    named_styles: List[NamedStyle] = field(default_factory=list)
    auto_filter_range: Optional[str] = None
    freeze_panes: Optional[str] = None


class FormattingExtractor:
    """Extrai formatação de planilhas Excel"""
    
    @staticmethod
    def extract_from_file(file_path: str, sheet_name: str = None) -> SheetFormatting:
        """
        Extrai toda a formatação de uma planilha Excel
        
        Args:
            file_path: Caminho para o arquivo Excel
            sheet_name: Nome da planilha (None para a primeira)
            
        Returns:
            SheetFormatting com toda a formatação extraída
        """
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=False)
            
            if sheet_name is None:
                worksheet = workbook.active
                sheet_name = worksheet.title
            else:
                worksheet = workbook[sheet_name]
            
            # Extrair formatação de células
            cell_formats = FormattingExtractor._extract_cell_formats(worksheet)
            
            # Extrair formatação de colunas
            column_formats = FormattingExtractor._extract_column_formats(worksheet)
            
            # Extrair estilos nomeados
            named_styles = FormattingExtractor._extract_named_styles(workbook)
            
            # Extrair configurações da planilha
            auto_filter_range = FormattingExtractor._extract_auto_filter(worksheet)
            freeze_panes = FormattingExtractor._extract_freeze_panes(worksheet)
            
            return SheetFormatting(
                name=sheet_name,
                cell_formats=cell_formats,
                column_formats=column_formats,
                named_styles=named_styles,
                auto_filter_range=auto_filter_range,
                freeze_panes=freeze_panes
            )
            
        except Exception as e:
            print(f"Erro ao extrair formatação de {file_path}: {e}")
            return SheetFormatting(
                name=sheet_name or "Unknown"
            )
    
    @staticmethod
    def _extract_cell_formats(worksheet) -> Dict[str, CellFormatting]:
        """Extrai formatação de todas as células com dados"""
        cell_formats = {}
        
        try:
            # Limitar a área de extração para evitar problemas de performance
            max_row = min(worksheet.max_row, 1000) if worksheet.max_row else 100
            max_col = min(worksheet.max_column, 50) if worksheet.max_column else 20
            
            for row in range(1, max_row + 1):
                for col in range(1, max_col + 1):
                    cell = worksheet.cell(row=row, column=col)
                    
                    # Verificar se a célula tem valor ou formatação
                    if cell.value is not None or (hasattr(cell, 'has_style') and cell.has_style):
                        cell_address = cell.coordinate
                        
                        try:
                            # Criar formatação de forma segura, evitando deepcopy
                            cell_format = CellFormatting(
                                font=FormattingExtractor._safe_copy_font(cell.font) if cell.font else None,
                                fill=FormattingExtractor._safe_copy_fill(cell.fill) if cell.fill else None,
                                border=FormattingExtractor._safe_copy_border(cell.border) if cell.border else None,
                                alignment=FormattingExtractor._safe_copy_alignment(cell.alignment) if cell.alignment else None,
                                number_format=str(cell.number_format) if cell.number_format else None,
                                formula=str(cell.value) if hasattr(cell, 'data_type') and cell.data_type == 'f' else None
                            )
                            
                            cell_formats[cell_address] = cell_format
                            
                        except Exception as e:
                            print(f"Erro ao extrair formatação da célula {cell_address}: {e}")
                            continue
                            
        except Exception as e:
            print(f"Erro geral ao extrair formatação de células: {e}")
        
        return cell_formats
    
    @staticmethod
    def _extract_column_formats(worksheet) -> Dict[str, ColumnFormatting]:
        """Extrai formatação de colunas"""
        column_formats = {}
        
        for col_letter in worksheet.column_dimensions:
            col_dim = worksheet.column_dimensions[col_letter]
            
            column_format = ColumnFormatting(
                width=col_dim.width,
                hidden=col_dim.hidden,
                auto_filter=False  # Será definido separadamente
            )
            
            column_formats[col_letter] = column_format
        
        return column_formats
    
    @staticmethod
    def _extract_named_styles(workbook) -> List[NamedStyle]:
        """Extrai estilos nomeados do workbook"""
        try:
            # Retorna lista vazia por enquanto para evitar problemas
            # Em versões futuras, pode ser implementado de forma segura
            return []
        except Exception as e:
            print(f"Erro ao extrair estilos nomeados: {e}")
            return []
    
    @staticmethod
    def _extract_auto_filter(worksheet) -> Optional[str]:
        """Extrai configuração de auto filtro"""
        if worksheet.auto_filter and worksheet.auto_filter.ref:
            return worksheet.auto_filter.ref
        return None
    
    @staticmethod
    def _extract_freeze_panes(worksheet) -> Optional[str]:
        """Extrai configuração de painéis congelados"""
        try:
            if hasattr(worksheet, 'freeze_panes') and worksheet.freeze_panes:
                return str(worksheet.freeze_panes)
        except Exception as e:
            print(f"Erro ao extrair freeze panes: {e}")
        return None
    
    @staticmethod
    def _safe_copy_font(font):
        """Cria uma cópia segura do objeto Font"""
        try:
            if font:
                return copy.copy(font)
        except Exception as e:
            print(f"Erro ao copiar fonte: {e}")
        return None
    
    @staticmethod
    def _safe_copy_fill(fill):
        """Cria uma cópia segura do objeto PatternFill"""
        try:
            if fill and hasattr(fill, 'fill_type') and fill.fill_type:
                return PatternFill(
                    fill_type=getattr(fill, 'fill_type', None),
                    start_color=getattr(fill, 'start_color', None),
                    end_color=getattr(fill, 'end_color', None)
                )
        except Exception as e:
            print(f"Erro ao copiar preenchimento: {e}")
        return None
    
    @staticmethod
    def _safe_copy_border(border):
        """Cria uma cópia segura do objeto Border"""
        try:
            if border:
                return Border(
                    left=getattr(border, 'left', None),
                    right=getattr(border, 'right', None),
                    top=getattr(border, 'top', None),
                    bottom=getattr(border, 'bottom', None)
                )
        except Exception as e:
            print(f"Erro ao copiar borda: {e}")
        return None
    
    @staticmethod
    def _safe_copy_alignment(alignment):
        """Cria uma cópia segura do objeto Alignment"""
        try:
            if alignment:
                return Alignment(
                    horizontal=getattr(alignment, 'horizontal', None),
                    vertical=getattr(alignment, 'vertical', None),
                    wrap_text=getattr(alignment, 'wrap_text', None)
                )
        except Exception as e:
            print(f"Erro ao copiar alinhamento: {e}")
        return None


class FormattingApplier:
    """Aplica formatação em planilhas Excel"""
    
    @staticmethod
    def apply_to_worksheet(worksheet, formatting: SheetFormatting, 
                          target_range: str = None) -> bool:
        """
        Aplica formatação a uma planilha
        
        Args:
            worksheet: Planilha do openpyxl
            formatting: Formatação a ser aplicada
            target_range: Faixa específica para aplicar (None para toda a planilha)
            
        Returns:
            True se aplicado com sucesso
        """
        try:
            # Aplicar formatação de células
            FormattingApplier._apply_cell_formats(worksheet, formatting.cell_formats, target_range)
            
            # Aplicar formatação de colunas
            FormattingApplier._apply_column_formats(worksheet, formatting.column_formats)
            
            # Aplicar auto filtro
            if formatting.auto_filter_range:
                FormattingApplier._apply_auto_filter(worksheet, formatting.auto_filter_range)
            
            # Aplicar painéis congelados
            if formatting.freeze_panes:
                FormattingApplier._apply_freeze_panes(worksheet, formatting.freeze_panes)
            
            return True
            
        except Exception as e:
            print(f"Erro ao aplicar formatação: {e}")
            return False
    
    @staticmethod
    def _apply_cell_formats(worksheet, cell_formats: Dict[str, CellFormatting], 
                           target_range: str = None):
        """Aplica formatação de células"""
        for cell_address, cell_format in cell_formats.items():
            try:
                cell = worksheet[cell_address]
                
                if cell_format.font:
                    cell.font = cell_format.font
                if cell_format.fill:
                    cell.fill = cell_format.fill
                if cell_format.border:
                    cell.border = cell_format.border
                if cell_format.alignment:
                    cell.alignment = cell_format.alignment
                if cell_format.number_format:
                    cell.number_format = cell_format.number_format
                if cell_format.formula:
                    cell.formula = cell_format.formula
                    
            except Exception as e:
                print(f"Erro ao aplicar formatação na célula {cell_address}: {e}")
    
    @staticmethod
    def _apply_column_formats(worksheet, column_formats: Dict[str, ColumnFormatting]):
        """Aplica formatação de colunas"""
        for col_letter, col_format in column_formats.items():
            try:
                col_dim = worksheet.column_dimensions[col_letter]
                
                if col_format.width:
                    col_dim.width = col_format.width
                if col_format.hidden:
                    col_dim.hidden = col_format.hidden
                    
            except Exception as e:
                print(f"Erro ao aplicar formatação na coluna {col_letter}: {e}")
    
    @staticmethod
    def _apply_auto_filter(worksheet, auto_filter_range: str):
        """Aplica auto filtro"""
        try:
            worksheet.auto_filter.ref = auto_filter_range
        except Exception as e:
            print(f"Erro ao aplicar auto filtro: {e}")
    
    @staticmethod
    def _apply_freeze_panes(worksheet, freeze_panes: str):
        """Aplica painéis congelados"""
        try:
            worksheet.freeze_panes = freeze_panes
        except Exception as e:
            print(f"Erro ao aplicar painéis congelados: {e}")


class FormattingMerger:
    """Gerencia mesclagem e conflitos de formatação"""
    
    @staticmethod
    def merge_formats(master_format: SheetFormatting, 
                     subordinate_formats: List[SheetFormatting],
                     priority: FormattingPriority = FormattingPriority.SUBORDINATE) -> SheetFormatting:
        """
        Mescla formatações de múltiplas planilhas
        
        Args:
            master_format: Formatação da planilha mestre
            subordinate_formats: Lista de formatações das subordinadas
            priority: Estratégia de prioridade para conflitos
            
        Returns:
            SheetFormatting mesclada
        """
        if priority == FormattingPriority.MASTER:
            return FormattingMerger._merge_with_master_priority(master_format, subordinate_formats)
        elif priority == FormattingPriority.SUBORDINATE:
            return FormattingMerger._merge_with_subordinate_priority(master_format, subordinate_formats)
        else:  # MERGE
            return FormattingMerger._merge_intelligently(master_format, subordinate_formats)
    
    @staticmethod
    def _merge_with_master_priority(master_format: SheetFormatting, 
                                   subordinate_formats: List[SheetFormatting]) -> SheetFormatting:
        """Mescla dando prioridade à formatação mestre"""
        merged = copy.deepcopy(master_format)
        
        # Adicionar formatações das subordinadas que não conflitam
        for sub_format in subordinate_formats:
            for cell_addr, cell_format in sub_format.cell_formats.items():
                if cell_addr not in merged.cell_formats:
                    merged.cell_formats[cell_addr] = cell_format
            
            for col_letter, col_format in sub_format.column_formats.items():
                if col_letter not in merged.column_formats:
                    merged.column_formats[col_letter] = col_format
        
        return merged
    
    @staticmethod
    def _merge_with_subordinate_priority(master_format: SheetFormatting, 
                                        subordinate_formats: List[SheetFormatting]) -> SheetFormatting:
        """Mescla dando prioridade às formatações subordinadas"""
        merged = copy.deepcopy(master_format)
        
        # Sobrescrever com formatações das subordinadas
        for sub_format in subordinate_formats:
            merged.cell_formats.update(sub_format.cell_formats)
            merged.column_formats.update(sub_format.column_formats)
            
            # Usar auto filtro da primeira subordinada que tiver
            if sub_format.auto_filter_range and not merged.auto_filter_range:
                merged.auto_filter_range = sub_format.auto_filter_range
        
        return merged
    
    @staticmethod
    def _merge_intelligently(master_format: SheetFormatting, 
                            subordinate_formats: List[SheetFormatting]) -> SheetFormatting:
        """Mescla de forma inteligente, preservando o melhor de cada formatação"""
        merged = copy.deepcopy(master_format)
        
        # Implementar lógica inteligente de mesclagem
        # Manter formatação de cabeçalho da mestre, dados das subordinadas
        for sub_format in subordinate_formats:
            # Aplicar formatação de dados (linhas > 1)
            for cell_addr, cell_format in sub_format.cell_formats.items():
                try:
                    # Extrair linha do endereço da célula
                    row_num = int(''.join(filter(str.isdigit, cell_addr)))
                    if row_num > 1:  # Não sobrescrever cabeçalho
                        merged.cell_formats[cell_addr] = cell_format
                except ValueError:
                    continue
            
            # Mesclar formatação de colunas
            merged.column_formats.update(sub_format.column_formats)
        
        return merged


class FormattingValidator:
    """Valida compatibilidade de formatação"""
    
    @staticmethod
    def validate_compatibility(formats: List[SheetFormatting]) -> Dict[str, List[str]]:
        """
        Valida se as formatações são compatíveis entre si
        
        Args:
            formats: Lista de formatações para validar
            
        Returns:
            Dicionário com warnings e erros encontrados
        """
        issues = {"warnings": [], "errors": []}
        
        if len(formats) < 2:
            return issues
        
        # Verificar compatibilidade de colunas
        FormattingValidator._check_column_compatibility(formats, issues)
        
        # Verificar conflitos de formatação
        FormattingValidator._check_formatting_conflicts(formats, issues)
        
        return issues
    
    @staticmethod
    def _check_column_compatibility(formats: List[SheetFormatting], issues: Dict[str, List[str]]):
        """Verifica compatibilidade de colunas"""
        all_columns = set()
        for fmt in formats:
            all_columns.update(fmt.column_formats.keys())
        
        for col in all_columns:
            widths = []
            for fmt in formats:
                if col in fmt.column_formats and fmt.column_formats[col].width:
                    widths.append(fmt.column_formats[col].width)
            
            if len(set(widths)) > 1:
                issues["warnings"].append(f"Coluna {col} tem larguras diferentes: {widths}")
    
    @staticmethod
    def _check_formatting_conflicts(formats: List[SheetFormatting], issues: Dict[str, List[str]]):
        """Verifica conflitos de formatação"""
        # Implementar verificações específicas conforme necessário
        pass


class FormattingService:
    """Serviço principal para gerenciamento de formatação"""
    
    def __init__(self):
        self.extractor = FormattingExtractor()
        self.applier = FormattingApplier()
        self.merger = FormattingMerger()
        self.validator = FormattingValidator()
    
    def process_formatting(self, master_file: str, subordinate_files: List[str], strategy: str = 'MERGE') -> bool:
        """
        Processa e aplica a formatação das planilhas subordinadas na mestre.
        """
        print(f"[FormattingService] Iniciando processamento de formatação para {len(subordinate_files)} arquivos.")
        if not os.path.exists(master_file):
            print(f"[FormattingService] ERRO: Arquivo mestre não encontrado em '{master_file}'")
            return False

        try:
            # Carregar a planilha mestre
            wb_master = openpyxl.load_workbook(master_file)
            ws_master = wb_master.active

            # Extrair a formatação de todas as planilhas subordinadas
            subordinate_formats = []
            for sub_file in subordinate_files:
                if os.path.exists(sub_file):
                    print(f"[FormattingService] Extraindo formatação de: {os.path.basename(sub_file)}")
                    sub_format = self.extractor.extract_from_file(sub_file)
                    subordinate_formats.append(sub_format)
                else:
                    print(f"[FormattingService] AVISO: Arquivo subordinado não encontrado: {sub_file}")

            if not subordinate_formats:
                print("[FormattingService] Nenhuma formatação de arquivo subordinado para aplicar.")
                return True

            # Lógica de mesclagem de formatação
            # Usar a formatação da primeira planilha subordinada como base para simplicidade
            # Em um cenário real, uma estratégia de mesclagem mais complexa seria necessária
            base_format = subordinate_formats[0]

            # Aplicar formatação de colunas
            print("[FormattingService] Aplicando formatação de colunas...")
            for col_letter, col_format in base_format.column_formats.items():
                dim = ws_master.column_dimensions[col_letter]
                if col_format.width:
                    dim.width = col_format.width
                if col_format.hidden:
                    dim.hidden = col_format.hidden

            # Aplicar formatação de células (cabeçalho)
            # Focar na primeira linha para evitar problemas de performance
            print("[FormattingService] Aplicando formatação de cabeçalho...")
            for row_idx in range(1, 2): # Apenas a primeira linha (cabeçalho)
                for col_idx in range(1, ws_master.max_column + 1):
                    cell_addr = f"{openpyxl.utils.get_column_letter(col_idx)}{row_idx}"
                    if cell_addr in base_format.cell_formats:
                        cell_format = base_format.cell_formats[cell_addr]
                        self._copy_cell_style(ws_master[cell_addr], cell_format)

            # Salvar as alterações
            wb_master.save(master_file)
            print(f"[FormattingService] Formatação aplicada e salva com sucesso em {master_file}")
            return True

        except Exception as e:
            print(f"[FormattingService] ERRO CRÍTICO ao processar formatação: {e}")
            import traceback
            traceback.print_exc()
            return False

    def _copy_cell_style(self, target_cell, cell_format: CellFormatting):
        """Copia o estilo de uma formatação de célula para uma célula de destino."""
        if cell_format.font:
            target_cell.font = copy.copy(cell_format.font)
        if cell_format.fill:
            target_cell.fill = copy.copy(cell_format.fill)
        if cell_format.border:
            target_cell.border = copy.copy(cell_format.border)
        if cell_format.alignment:
            target_cell.alignment = copy.copy(cell_format.alignment)
        if cell_format.number_format:
            target_cell.number_format = cell_format.number_format
    
    def extract_formatting_summary(self, file_path: str) -> Dict[str, Any]:
        """
        Extrai um resumo da formatação de um arquivo
        
        Args:
            file_path: Caminho do arquivo
            
        Returns:
            Dicionário com resumo da formatação
        """
        try:
            formatting = self.extractor.extract_from_file(file_path)
            
            summary = {
                "sheet_name": formatting.name,
                "total_formatted_cells": len(formatting.cell_formats),
                "total_formatted_columns": len(formatting.column_formats),
                "has_auto_filter": formatting.auto_filter_range is not None,
                "has_freeze_panes": formatting.freeze_panes is not None,
                "named_styles_count": len(formatting.named_styles)
            }
            
            return summary
            
        except Exception as e:
            print(f"Erro ao extrair resumo de formatação: {e}")
            return {}