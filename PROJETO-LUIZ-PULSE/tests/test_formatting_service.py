import unittest
import os
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill

# Add the project root to the Python path
project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
sys.path.insert(0, project_root)

from services.formatting_service import FormattingService

class TestFormattingService(unittest.TestCase):

    def setUp(self):
        self.service = FormattingService()
        self.test_dir = "test_data"
        os.makedirs(self.test_dir, exist_ok=True)

        # Create a dummy master file
        self.master_file = os.path.join(self.test_dir, "master.xlsx")
        wb_master = openpyxl.Workbook()
        ws_master = wb_master.active
        ws_master["A1"] = "Header 1"
        ws_master["B1"] = "Header 2"
        wb_master.save(self.master_file)

        # Create a dummy subordinate file with formatting
        self.subordinate_file = os.path.join(self.test_dir, "subordinate.xlsx")
        wb_sub = openpyxl.Workbook()
        ws_sub = wb_sub.active
        ws_sub["A1"] = "Sub Header 1"
        ws_sub["A1"].font = Font(bold=True, color="FF0000")
        ws_sub["B1"] = "Sub Header 2"
        ws_sub["B1"].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        ws_sub.column_dimensions['A'].width = 20
        wb_sub.save(self.subordinate_file)

    def tearDown(self):
        import shutil
        shutil.rmtree(self.test_dir)

    def test_process_formatting(self):
        # Run the formatting service
        success = self.service.process_formatting(self.master_file, [self.subordinate_file])
        self.assertTrue(success)

        # Check if the formatting was applied correctly
        wb_master = openpyxl.load_workbook(self.master_file)
        ws_master = wb_master.active

        # Check font
        self.assertTrue(ws_master["A1"].font.bold)
        self.assertEqual(ws_master["A1"].font.color.rgb, "00FF0000")

        # Check fill
        self.assertEqual(ws_master["B1"].fill.start_color.rgb, "00FFFF00")

        # Check column width
        self.assertEqual(ws_master.column_dimensions['A'].width, 20)

if __name__ == "__main__":
    unittest.main()
